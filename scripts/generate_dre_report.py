#!/usr/bin/env python3
"""
Gera o relatório "One Page" (Consolidado / Varejo / Atacado) a partir do
arquivo "DRE Aberta Mensal" exportado da planilha DRE Banco (Google Sheets).

Uso:
    python3 scripts/generate_dre_report.py \
        --file data/DRE_Aberta_Mensal.xlsx \
        --year 2026 --months 1-6 \
        --compare-year 2025 --compare-months 1-6 \
        --label "6 M/2026" \
        --slide-number 3 \
        --output reports/one-page-6M-2026.html

Como exportar a planilha do Google Sheets antes de rodar o script:
    Arquivo > Fazer download > Microsoft Excel (.xlsx)
    Salve em data/DRE_Aberta_Mensal.xlsx (ou outro caminho e use --file).

O arquivo precisa ter uma aba por ano (ex: "2025", "2026", ...) com as colunas:
    Mês Ano | Loja | Canal | DRE | -<ano> | Orçado -<ano>
"""

import argparse
import sys
from pathlib import Path

import openpyxl

TOP_LEVEL_METRICS = {
    "receita_bruta": "Receita Bruta",
    "receita_liquida": "Receita Líquida",
    "lucro_bruto_ajustado": "Lucro Bruto Ajustado",
    "ebitda": "EBITDA",
}

# Canal(is) da aba DRE Banco que compõem cada coluna do relatório.
# Consolidado = soma de todos os canais presentes na planilha.
SEGMENTS = {
    "Consolidado": None,
    "Varejo": {"VAREJO"},
    "Atacado": {"ATACADO"},
}

# Nota exibida no relatório: a coluna Atacado do One Page oficial usa a base
# "Lynkz Gerencial Consolidado", que pode diferir da aba ATACADO da DRE Banco
# (contábil). Ajuste/remova esta nota se a fonte do Atacado mudar.
ATACADO_NOTE = "Atacado: Lynkz Gerencial Consolidado"


def parse_months(spec):
    """Aceita '1-6' ou '1,2,3,4,5,6' e devolve um set de inteiros 1-12."""
    months = set()
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-")
            months.update(range(int(start), int(end) + 1))
        else:
            months.add(int(part))
    return months


def load_rows(workbook_path, year):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet_name = str(year)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Aba '{sheet_name}' não encontrada no arquivo. Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt, loja, canal, dre, atual, orcado = (row + (None,) * 6)[:6]
        if dt is None or dre is None:
            continue
        rows.append((dt, loja, canal, dre, atual or 0.0))
    return rows


def sum_metric(rows, months, metric_label, canais):
    total = 0.0
    for dt, _loja, canal, dre, valor in rows:
        if dt.month not in months:
            continue
        if canais is not None and canal not in canais:
            continue
        if dre.startswith("\xa0"):
            continue  # ignora linhas de detalhe (contas contábeis individuais)
        if dre.strip() != metric_label:
            continue
        total += valor
    return total


def compute_segment_metrics(rows, months, canais):
    values = {
        key: sum_metric(rows, months, label, canais)
        for key, label in TOP_LEVEL_METRICS.items()
    }
    receita_liquida = values["receita_liquida"]
    values["margem_bruta"] = (
        values["lucro_bruto_ajustado"] / receita_liquida if receita_liquida else 0.0
    )
    values["margem_ebitda"] = (
        values["ebitda"] / receita_liquida if receita_liquida else 0.0
    )
    return values


def variance(curr, prev):
    delta = curr - prev
    pct = (delta / abs(prev)) if prev else None
    return delta, pct


def fmt_money(value_reais, decimals=1):
    """Formata em milhões, estilo pt-BR (vírgula decimal). Ex: R$ 91,3MM."""
    mm = value_reais / 1_000_000
    sign = "-" if mm < 0 else ""
    text = f"{abs(mm):.{decimals}f}".replace(".", ",")
    return f"{sign}R$ {text}MM"

def fmt_money_delta(value_reais, decimals=1):
    """Delta em MM, ou em mil (K) quando muito pequeno, com sinal explícito."""
    sign = "+" if value_reais >= 0 else "-"
    abs_value = abs(value_reais)
    if abs_value < 50_000:
        text = f"{abs_value / 1_000:.1f}".replace(".", ",")
        return f"{sign}R$ {text}K"
    text = f"{abs_value / 1_000_000:.{decimals}f}".replace(".", ",")
    return f"{sign}R$ {text}MM"


def fmt_pct(value, decimals=1):
    sign = "+" if value >= 0 else "-"
    text = f"{abs(value) * 100:.{decimals}f}".replace(".", ",")
    return f"{sign}{text}%"


def fmt_pp(value, decimals=1):
    sign = "+" if value >= 0 else "-"
    text = f"{abs(value) * 100:.{decimals}f}".replace(".", ",")
    return f"{sign}{text} p.p."


def color_for(value):
    if value is None:
        return "neutral"
    if value > 1e-9:
        return "positive"
    if value < -1e-9:
        return "negative"
    return "neutral"


def build_segment_html(name, curr, prev, compare_label, show_note=False):
    rb_delta, rb_pct = variance(curr["receita_bruta"], prev["receita_bruta"])
    lba_delta, lba_pct = variance(curr["lucro_bruto_ajustado"], prev["lucro_bruto_ajustado"])
    mb_pp = curr["margem_bruta"] - prev["margem_bruta"]
    ebitda_delta, ebitda_pct = variance(curr["ebitda"], prev["ebitda"])
    me_pp = curr["margem_ebitda"] - prev["margem_ebitda"]

    def pct_line(pct, delta):
        pct_txt = fmt_pct(pct) if pct is not None else "n/a"
        return f"{pct_txt} | {fmt_money_delta(delta)} vs {compare_label}"

    note_html = ""
    if show_note:
        note_html = f'<div class="segment-note">{ATACADO_NOTE}</div>'

    return f"""
    <div class="segment">
      <div class="segment-header">{name}</div>
      <div class="segment-body">
        <div class="metric">
          <div class="metric-title">RECEITA BRUTA: {fmt_money(curr['receita_bruta'])}</div>
          <div class="metric-sub {color_for(rb_pct)}">{pct_line(rb_pct, rb_delta)}</div>
        </div>
        <div class="metric">
          <div class="metric-title">LUCRO BRUTO AJUSTADO: {fmt_money(curr['lucro_bruto_ajustado'])}</div>
          <div class="metric-sub {color_for(lba_pct)}">{pct_line(lba_pct, lba_delta)}</div>
        </div>
        <div class="metric">
          <div class="metric-title">MARGEM BRUTA: {fmt_pct(curr['margem_bruta']).lstrip('+')}</div>
          <div class="metric-sub {color_for(mb_pp)}">{fmt_pp(mb_pp)} vs {compare_label}</div>
        </div>
        <div class="metric">
          <div class="metric-title">EBITDA: {fmt_money(curr['ebitda'])}</div>
          <div class="metric-sub {color_for(ebitda_pct)}">{pct_line(ebitda_pct, ebitda_delta)}</div>
        </div>
        <div class="metric">
          <div class="metric-title">MARGEM EBITDA: {fmt_pct(curr['margem_ebitda']).lstrip('+')}</div>
          <div class="metric-sub {color_for(me_pp)}">{fmt_pp(me_pp)} vs {compare_label}</div>
        </div>
      </div>
      {note_html}
    </div>
    """


PAGE_TEMPLATE = """<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>One Page - {label}</title>
<style>
  body {{
    font-family: Calibri, Arial, sans-serif;
    margin: 0;
    padding: 24px;
    background: #ffffff;
    color: #1a1a1a;
  }}
  .page {{
    max-width: 1200px;
    margin: 0 auto;
    border: 1px solid #d9e5dd;
  }}
  .title-bar {{
    display: flex;
    align-items: center;
    gap: 20px;
    background: linear-gradient(90deg, #1f5c3f 0%, #3f8f68 55%, #bfe3cf 100%);
    color: #ffffff;
    padding: 14px 24px;
  }}
  .title-bar .badge {{
    background: #17442e;
    width: 40px;
    height: 40px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: bold;
    font-size: 20px;
    border-radius: 4px;
  }}
  .title-bar .title {{
    font-size: 24px;
    font-weight: bold;
    text-decoration: underline;
  }}
  .period-bar {{
    background: #eef7f0;
    padding: 10px 24px;
    font-size: 20px;
    font-weight: bold;
    color: #1a1a1a;
  }}
  .segments {{
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 20px;
    padding: 24px;
  }}
  .segment-header {{
    background: #b7e3c9;
    color: #1a1a1a;
    text-align: center;
    font-weight: bold;
    padding: 8px;
    border-radius: 6px;
    margin-bottom: 12px;
  }}
  .segment-body {{
    background: #f4f4f4;
    border-radius: 6px;
    padding: 18px 16px;
    display: flex;
    flex-direction: column;
    gap: 16px;
  }}
  .metric {{
    text-align: center;
  }}
  .metric-title {{
    font-weight: bold;
    font-size: 14px;
  }}
  .metric-sub {{
    font-size: 13px;
    margin-top: 2px;
  }}
  .positive {{ color: #2e7d32; }}
  .negative {{ color: #c00000; }}
  .neutral  {{ color: #1a1a1a; }}
  .segment-note {{
    margin-top: 10px;
    border: 1px solid #999;
    border-radius: 4px;
    padding: 4px 8px;
    font-size: 11px;
    font-weight: bold;
    background: #ffffff;
    text-align: center;
  }}
  .generated-at {{
    text-align: right;
    font-size: 10px;
    color: #888;
    padding: 0 24px 16px;
  }}
</style>
</head>
<body>
<div class="page">
  <div class="title-bar">
    <div class="badge">{slide_number}</div>
    <div class="title">One Page</div>
  </div>
  <div class="period-bar">{label}</div>
  <div class="segments">
    {consolidado_html}
    {varejo_html}
    {atacado_html}
  </div>
  <div class="generated-at">Gerado automaticamente a partir da planilha DRE Banco (Google Sheets)</div>
</div>
</body>
</html>
"""


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--file", required=True, help="Caminho do .xlsx exportado da DRE Banco")
    parser.add_argument("--year", type=int, required=True, help="Ano do período atual (ex: 2026)")
    parser.add_argument("--months", required=True, help="Meses do período atual, ex: 1-6")
    parser.add_argument("--compare-year", type=int, required=True, help="Ano do período de comparação (ex: 2025)")
    parser.add_argument("--compare-months", required=True, help="Meses do período de comparação, ex: 1-6")
    parser.add_argument("--label", required=True, help="Rótulo exibido no relatório, ex: '6 M/2026'")
    parser.add_argument("--compare-label", default=None, help="Rótulo do período de comparação nas variações (padrão: 6M/<compare-year>)")
    parser.add_argument("--slide-number", default="3", help="Número exibido no badge do cabeçalho")
    parser.add_argument("--output", required=True, help="Caminho do .html de saída")
    args = parser.parse_args()

    months = parse_months(args.months)
    compare_months = parse_months(args.compare_months)
    compare_label = args.compare_label or f"6M/{args.compare_year}"

    curr_rows = load_rows(args.file, args.year)
    prev_rows = (
        load_rows(args.file, args.compare_year)
        if args.compare_year != args.year
        else curr_rows
    )

    sections = {}
    for name, canais in SEGMENTS.items():
        curr = compute_segment_metrics(curr_rows, months, canais)
        prev = compute_segment_metrics(prev_rows, compare_months, canais)
        sections[name] = build_segment_html(
            name, curr, prev, compare_label, show_note=(name == "Atacado")
        )

    html = PAGE_TEMPLATE.format(
        label=args.label,
        slide_number=args.slide_number,
        consolidado_html=sections["Consolidado"],
        varejo_html=sections["Varejo"],
        atacado_html=sections["Atacado"],
    )

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    print(f"Relatório gerado em: {output_path}")


if __name__ == "__main__":
    sys.exit(main())
